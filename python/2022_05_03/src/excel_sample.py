# --*--coding:utf-8--*--
# from operator import truediv, truth
import openpyxl
import random

# sampleを使うためのimport
# from openpyxl import Workbook
# from openpyxl.utils import get_column_letter


# def sample():
#     wb = Workbook()
#     dest_filename = 'empty_book.xlsx'
#     ws1 = wb.active
#     ws1.title = "range names"
#     for row in range(1, 40):
#         ws1.append(range(600))
#     ws2 = wb.create_sheet(title="Pi")
#     ws2['F5'] = 3.14
#     ws3 = wb.create_sheet(title="Data")
#     for row in range(10, 20):
#         for col in range(27, 54):
#             _ = ws3.cell(column=col, row=row,
#                          value="{0}".format(get_column_letter(col)))
#     print(ws3['AA10'].value)
#     wb.save(filename=dest_filename)

def makeTickTackToe():
    ws = openpyxl.Workbook()
    sheet1 = ws.create_sheet(title='Sheet1')
    for row in range(1, 4):
        for columns in range(1, 4):
            sheet1.cell(row=row, column=columns, value='●')
    Sheet2 = ws.create_sheet(title='Sheet2')
    for row in range(1, 4):
        for columns in range(1, 4):
            if row == 2 and columns == 2:
                Sheet2.cell(row=row, column=columns, value='×')
                continue
            Sheet2.cell(row=row, column=columns, value='●')
    ws.save('ticktacktoe.xlsx')


def makeWorkBook():
    wb = openpyxl.Workbook()
    print('make sheet')
    wb.create_sheet()
    print(wb.sheetnames)
    sheet = wb[wb.sheetnames[0]]
    for count1 in range(1, 4):
        for count2 in range(1, 4):
            sheet.cell(row=count1, column=count2, value='●')
    wb.save('aaa.xlsx')


def checkDiff():
    ws = openpyxl.load_workbook('ticktacktoe.xlsx', data_only=True)
    sheet1 = ws[ws.sheetnames[1]]
    sheet2 = ws[ws.sheetnames[2]]
    array1 = [i for i in sheet1.values]
    array2 = [i for i in sheet2.values]
    print(array1)
    print(array2)
    length = len(array1)
    for i in range(length):
        if array1[i] == array2[i]:
            print(f'{i+1}は同じ値を持ちます')
        else:
            for s in range(len(array1[i])):
                if array1[i][s] != array2[i][s]:
                    print(f'{i+1}行{s+1}列の値が異なる')
                else:
                    continue


def main():
    # workBook = openpyxl.load_workbook('../text_file/work.xlsx', data_only=True)
    # # print(workBook.sheetnames)
    # sheetList = workBook.sheetnames
    # sheet0 = workBook[sheetList[0]]
    # print(type(sheet0))
    # for row in sheet0.values:
    #     print(row)
    makeWorkBook()
    # sample()
    # makeTickTackToe()
    # checkDiff()


if __name__ == '__main__':
    main()
