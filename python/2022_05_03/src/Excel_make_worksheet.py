# --*-- coding:utf-8--*--
from asyncio.windows_events import NULL
import string
import openpyxl
from openpyxl.formula.translate import Translator
import glob
import sys


def makeDate(num=100) -> list:
    allList = []
    smallList = []
    for i in range(num):
        if i % 3 == 0 and i != 0:
            allList.append(smallList)
            smallList = []
            smallList.append(i)
        else:
            smallList.append(i)
    return allList


def execute1():
    # create work sheet
    ws = openpyxl.Workbook()
    print(ws.sheetnames)
    retData = makeDate()
    # openpyxlではappendメソッドを使用することによって、行を追加することができる
    for retDataElem in retData:
        ws[ws.sheetnames[0]].append(retDataElem)
    ws.save('../ExcelFile/2022_05_05.xlsx')
    ws.close()


def execute2():
    ws = openpyxl.load_workbook('../ExcelFile/2022_05_05.xlsx', data_only=True)
    print(ws.sheetnames)
    workSheet = ws[ws.sheetnames[0]]
    for rows in range(1, 34):
        translateCordinate = workSheet.cell(row=rows, column=4).coordinate
        fomulas = Translator(
            "=SUM(A1:C1)", origin="D1").translate_formula(translateCordinate)
        workSheet[translateCordinate] = fomulas
    ws.save('../ExcelFile/2022_05_05.xlsx')


def getCommandArg() -> string:
    print('command line args number is '+str(len(sys.argv)))
    if len(sys.argv) <= 1:
        print('grep Excel file target is current directory')
        return './*.xlsx'
    else:
        print('grep target directory is {}'.format(sys.argv[1]))
        return str(sys.argv[1])+'/*.xlsx'


def globExcelpath(path='./*') -> list:
    return glob.glob(path)


def searchWorkBook(workBook, searchKey) -> list:
    if workBook is None:
        print('method execute is finished')
        return
    sheetsList = workBook.sheetnames
    retSheetsList = []
    for sheetName in sheetsList:
        sheet = workBook[sheetName]
        for row in sheet.values:
            if searchKey in row:
                retSheetsList.append(sheetName)
                break
    if retSheetsList != []:
        return retSheetsList
    else:
        return NULL


if __name__ == '__main__':
    # execute1()
    # execute2()
    # grep処理の実装
    searchDirecotry = getCommandArg()
    targetExcelFiles = globExcelpath(searchDirecotry)
    print('target file '+','.join(targetExcelFiles))
    for targetExcelFile in targetExcelFiles:
        wb = openpyxl.load_workbook(targetExcelFile, data_only=True)
        searchKey = '1'
        searchKey = '×'
        retVal = searchWorkBook(wb, searchKey)
        if retVal != NULL:
            print('{0} is find {1} of {2}'.format(
                searchKey, targetExcelFile, ','.join(retVal)))
        else:
            # print('{0} is not find {1}'.format(searchKey, targetExcelFile))
            pass
