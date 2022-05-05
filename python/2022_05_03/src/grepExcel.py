import glob
from importlib.resources import path
from operator import truediv
import re
import openpyxl
import os


def lsdir(path='./'):
    return glob.glob(path)


def grep(searchKey):
    targetList = lsdir('../ExcelFile/*.xlsx')
    for target in targetList:
        target2 = target.replace('\\', '/')
        print(target2)
        # openpyxl.load_workbook(filename=target2, data_only=True)
        # with openpyxl.load_workbook(filename=str(target2), data_only=True) as rExcel:
        rExcel = openpyxl.load_workbook(filename=str(target2), data_only=True)
        sheetsList = rExcel.sheetnames
        for sheet in sheetsList:
            ws = rExcel[sheet]
            for row in ws.values:
                if searchKey in row:
                    print(f'{searchKey} is find in {target}:{sheet}')
        rExcel.close()


def main():
    grep('●')


if __name__ == '__main__':
    main()
