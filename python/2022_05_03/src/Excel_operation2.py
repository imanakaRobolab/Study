from openpyxl.styles import Border, Side
import openpyxl


border = Border(left=Side(border_style='thin', color='FF000000'))

ws = openpyxl.Workbook()
sheet = ws[ws.sheetnames[0]]


for i in sheet['A1:B4']:
    for cell in i:
        cell.value = 'a'
        cell.border = border

ws.save('aaa.xlsx')
ws.close()
